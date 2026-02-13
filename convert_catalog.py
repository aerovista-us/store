import json
import os
import re
import sys
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except Exception:
        s = _cell_str(v)
        s = re.sub(r"[^\d.\-]", "", s)
        try:
            return float(s) if s else 0.0
        except Exception:
            return 0.0


def _strip_html(s: str) -> str:
    if not s:
        return ""
    # Cheap tag stripper for export content.
    s = re.sub(r"<[^>]+>", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _extract_date_from_filename(path: str) -> Optional[str]:
    m = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(path))
    return m.group(1) if m else None


def _find_header_row(ws) -> Tuple[int, List[str]]:
    """
    Square export files often have a blank first row. Find the first row that contains
    'Reference Handle' and use it as the header row.
    """
    for r in range(1, 25):
        values = [c.value for c in next(ws.iter_rows(min_row=r, max_row=r))]
        if any(_cell_str(v).lower() == "reference handle" for v in values):
            headers = [_cell_str(v) for v in values]
            return r, headers
    raise RuntimeError("Unable to locate header row (expected 'Reference Handle').")


def _parse_items_sheet(xlsx_path: str) -> List[Dict[str, Any]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Items"] if "Items" in wb.sheetnames else wb[wb.sheetnames[0]]

    header_row, headers = _find_header_row(ws)
    rows = ws.iter_rows(min_row=header_row + 1, values_only=True)

    out_rows: List[Dict[str, Any]] = []
    for row in rows:
        if not row or all(v is None for v in row):
            continue
        d = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            if i < len(row):
                d[h] = row[i]
        if _cell_str(d.get("Reference Handle")):
            out_rows.append(d)
    return out_rows


def _name_and_color(item_name: str) -> Tuple[str, str]:
    s = (item_name or "").strip()
    m = re.match(r"^(.*)\(([^)]+)\)\s*$", s)
    if not m:
        return s, ""
    base = m.group(1).strip()
    color = m.group(2).strip()
    return base, color


def build_catalog(xlsx_path: str) -> Dict[str, Any]:
    rows = _parse_items_sheet(xlsx_path)

    products_by_id: Dict[str, Dict[str, Any]] = {}

    for r in rows:
        ref_handle = _cell_str(r.get("Reference Handle"))
        if not ref_handle.startswith("#"):
            # Skip non-handle rows (defensive).
            continue
        handle = ref_handle.lstrip("#").strip()
        if not handle:
            continue

        # Variation rows end with `--<option>` (e.g. `--s`).
        base_id = handle.split("--")[0].strip()
        if not base_id:
            continue

        token = _cell_str(r.get("Token"))
        sku = _cell_str(r.get("SKU"))
        variation_name = _cell_str(r.get("Variation Name"))
        opt1 = _cell_str(r.get("Option Value 1"))
        size = variation_name or opt1 or "One Size"

        item_name = _cell_str(r.get("Item Name"))
        base_name, color = _name_and_color(item_name)

        desc = _cell_str(r.get("Description"))
        desc_html = desc if ("<" in desc and ">" in desc) else (f"<p>{desc}</p>" if desc else "")
        desc_text = _strip_html(desc_html) if desc_html else _strip_html(desc)

        category = _cell_str(r.get("Reporting Category")) or _cell_str(r.get("Categories")) or ""
        visibility = _cell_str(r.get("Square Online Item Visibility")) or ""
        shipping_enabled = _cell_str(r.get("Shipping Enabled")) or ""

        price = _to_float(r.get("Price"))
        variant_price = price if price else _to_float(r.get("Online Sale Price"))

        p = products_by_id.get(base_id)
        if not p:
            p = {
                "id": base_id,
                "name": base_name or base_id,
                "color": color,
                "category": category,
                "price": float(variant_price or 0.0),
                "visibility": visibility or "visible",
                "shipping_enabled": shipping_enabled or "",
                "description_text": desc_text,
                "description_html": desc_html,
                "variants": [],
            }
            products_by_id[base_id] = p

        # Prefer first non-empty metadata across rows.
        if not p.get("color") and color:
            p["color"] = color
        if not p.get("category") and category:
            p["category"] = category
        if not p.get("description_text") and desc_text:
            p["description_text"] = desc_text
        if not p.get("description_html") and desc_html:
            p["description_html"] = desc_html
        if p.get("price", 0.0) == 0.0 and (variant_price or 0.0) != 0.0:
            p["price"] = float(variant_price)
        if p.get("visibility") in ("", None) and visibility:
            p["visibility"] = visibility
        if not p.get("shipping_enabled") and shipping_enabled:
            p["shipping_enabled"] = shipping_enabled

        # Add variant (token is the Square object id in exports; treat it as variation_id)
        variant: Dict[str, Any] = {
            "size": size,
            "sku": sku,
            "price": float(variant_price or 0.0),
        }
        if token:
            variant["variation_id"] = token
        p["variants"].append(variant)

    products = sorted(products_by_id.values(), key=lambda x: x.get("id", ""))
    return {
        "generated_from": f"Square catalog export Excel ({os.path.basename(xlsx_path)})",
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "count": len(products),
        "products": products,
    }


def main(argv: List[str]) -> int:
    xlsx_path = argv[1] if len(argv) > 1 else "1149XBNG8C8ZE_catalog-2026-02-11-0606.xlsx"
    if not os.path.exists(xlsx_path):
        print(f"ERROR: xlsx not found: {xlsx_path}", file=sys.stderr)
        return 2

    out = build_catalog(xlsx_path)
    if int(out.get("count") or 0) <= 0:
        print("ERROR: conversion produced 0 products; refusing to write outputs.", file=sys.stderr)
        return 3

    out_latest = "square_products_latest.json"
    date = _extract_date_from_filename(xlsx_path)
    out_dated = f"square_products_{date}.json" if date else None

    payload = json.dumps(out, indent=2, ensure_ascii=False)

    # Write atomically to prevent corrupting "latest" on a partial write.
    out_latest_tmp = "square_products_latest.tmp.json"
    out_dated_tmp = f"{out_dated}.tmp" if out_dated else None

    try:
        with open(out_latest_tmp, "w", encoding="utf-8") as f:
            f.write(payload + "\n")
        os.replace(out_latest_tmp, out_latest)

        if out_dated and out_dated_tmp:
            with open(out_dated_tmp, "w", encoding="utf-8") as f:
                f.write(payload + "\n")
            os.replace(out_dated_tmp, out_dated)
    finally:
        # Best-effort cleanup if something failed mid-write.
        for p in (out_latest_tmp, out_dated_tmp):
            if p and os.path.exists(p):
                try:
                    os.remove(p)
                except Exception:
                    pass

    print(f"Wrote {out_latest} (products={out['count']})")
    if out_dated:
        print(f"Wrote {out_dated}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
