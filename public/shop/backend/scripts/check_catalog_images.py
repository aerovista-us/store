#!/usr/bin/env python
"""
Quick helper to report which catalog rows in a Square export XLSX
are missing any image/thumbnail/photo value.

Usage (from repo root):

  python backend/scripts/check_catalog_images.py \
    1149XBNG8C8ZE_catalog-2026-03-11-2023.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path


def main(argv: list[str]) -> None:
    if len(argv) < 2:
        print("Usage: check_catalog_images.py <excel-file>", file=sys.stderr)
        raise SystemExit(1)

    xlsx_path = Path(argv[1]).expanduser()
    if not xlsx_path.is_file():
        print(f"File not found: {xlsx_path}", file=sys.stderr)
        raise SystemExit(1)

    try:
        import openpyxl  # type: ignore
    except ImportError:
        print("This script requires openpyxl. Install with:\n  pip install openpyxl", file=sys.stderr)
        raise SystemExit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Build header map
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: dict[str, int] = {}
    for idx, value in enumerate(header_row, start=1):
        key = (str(value).strip().lower() if value is not None else "")
        headers[key] = idx

    image_headers = [name for name in headers if any(k in name for k in ("image", "photo", "thumbnail", "thumb", "picture", "img", "asset"))]
    if not image_headers:
        print("No image/photo/thumbnail columns detected. Headers were:", file=sys.stderr)
        for k, idx in headers.items():
            print(f"  col {idx}: {repr(k)}", file=sys.stderr)
        raise SystemExit(1)

    name_col_idx = headers.get("name") or headers.get("item name") or 1
    image_col_indexes = [headers[h] for h in image_headers]

    missing: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name_val = row[name_col_idx - 1]
        has_image = False
        for col_idx in image_col_indexes:
            val = row[col_idx - 1]
            if val not in (None, "", "0"):
                has_image = True
                break
        if not has_image:
            missing.append(str(name_val or "").strip() or "(no name)")

    print(f"Image-related columns: {image_headers}")
    print(f"Total rows checked: {ws.max_row - 1}")
    print(f"Rows missing images: {len(missing)}")
    for name in missing:
        print(f"- {name}")


if __name__ == "__main__":
    main(sys.argv)

